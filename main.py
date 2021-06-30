import math
import openpyxl

from kivy.app import App
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.label import Label
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.properties import ObjectProperty

iv = ObjectProperty(None)
bdwt = ObjectProperty(None)


class MainScreen(ScreenManager):
    pass


class shape(Screen):
    pass


class bdwt(Screen):
    pass


class CglApp(App):
    def CgBdwt(self):

        x=float(self.root.ids.cgl_thk.text)
        y=float(self.root.ids.cgl_width.text)
        self.getbdwt(x,y)
    def getbdwt(self,x,y):
        wb = openpyxl.load_workbook('cgl_bdwt.xlsx')
        ws = wb['Sheet1']

        for i in range(1, ws.max_row):
            if (ws.cell(i, 1).value == x):
                if (ws.cell(i, 2).value == y):
                    self.root.ids.minbdwt.text = str(round((ws.cell(i, 3).value),3))
                    self.root.ids.avgbdwt.text = str(round((ws.cell(i, 4).value),3))
                    self.root.ids.maxbdwt.text = str(round((ws.cell(i, 5).value),3))
                    self.root.ids.warning.text = ''

    def build(self):
        return MainScreen()
if __name__=='__main__':
    CglApp().run()